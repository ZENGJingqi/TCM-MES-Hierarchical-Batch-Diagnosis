from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path("论文写作") / "supplementary_tables" / "Supplementary_Table_S2_layer_wise_batch_linkage_and_evidence.xlsx"


rows = [
    {
        "evidence_layer": "Jianwei Xiaoshi tablet quality testing",
        "linkage_path": "Jianwei Xiaoshi tablet quality testing -> Jianwei Xiaoshi tablet MES",
        "linked_scope": "908 tablet quality-testing records linked to tablet MES records",
        "analysis_unit": "Finished-product quality-testing record",
        "issue_endpoint": "0.8 g tablet disintegration time >10 min",
        "issue_count": "412/3375 records (12.21%) in the 0.8 g specification",
        "main_signal": "Temporally concentrated disintegration increase",
        "statistical_evidence": "Issue rates: 2024-08, 17/18 (94.44%); 2024-09, 146/176 (82.95%); 2026-01, 118/162 (72.84%); 2026-02, 122/144 (84.72%)",
        "interpretation": "The 0.8 g disintegration endpoint provided the issue-driven starting point for upstream diagnosis.",
    },
    {
        "evidence_layer": "Jianwei Xiaoshi extract-powder quality testing",
        "linkage_path": "Jianwei Xiaoshi extract powder -> Jianwei Xiaoshi tablet MES -> Jianwei Xiaoshi tablet quality testing",
        "linked_scope": "273 extract-powder batches linked forward to tablet MES; 933 downstream MES records traced back to extract powder; 804 tablet quality-testing records traced to extract-powder quality data",
        "analysis_unit": "Extract-powder batch",
        "issue_endpoint": "At least one linked tablet quality-testing record with disintegration time >10 min",
        "issue_count": "26/241 linked extract-powder batches (10.79%)",
        "main_signal": "Higher total ash and lower extractives",
        "statistical_evidence": "Total ash: Spearman rho = 0.311 with mean disintegration, P <0.001; OR per SD = 9.485. Extractives: rho = -0.210, P = 0.001; OR per SD = 0.136.",
        "interpretation": "The extract-powder layer narrowed the finished-product issue to a total-ash/extractives quality pattern.",
    },
    {
        "evidence_layer": "Chenpi quality testing",
        "linkage_path": "Chenpi -> Jianwei Xiaoshi extract powder -> Jianwei Xiaoshi tablet MES -> Jianwei Xiaoshi tablet quality testing",
        "linked_scope": "41 Chenpi testing batches linked forward; 578 downstream extract-powder batches traced back to Chenpi; 873 linked tablet quality-testing records in the Chenpi-to-tablet chain",
        "analysis_unit": "Chenpi batch / extract-powder batch mapped through traceability",
        "issue_endpoint": "Linked tablet quality-testing record with disintegration time >10 min",
        "issue_count": "28 extract-powder batches had at least one linked >10 min tablet record in the Chenpi-linked chain",
        "main_signal": "Lower Chenpi hesperidin and lower Chenpi moisture",
        "statistical_evidence": "Chenpi hesperidin vs linked issue rate: Spearman rho = -0.500, P <0.001. Chenpi moisture vs linked issue rate: rho = -0.434, P <0.001.",
        "interpretation": "Chenpi attributes provided upstream herbal-material evidence associated with the extract-powder and finished-product disintegration patterns.",
    },
    {
        "evidence_layer": "Chinese yam powder MES production records",
        "linkage_path": "Chinese yam powder -> Jianwei Xiaoshi tablet MES -> Jianwei Xiaoshi tablet quality testing",
        "linked_scope": "176 Chinese yam powder batches linked forward to tablet MES; 1,159 downstream MES records traced back to Chinese yam powder; 839 tablet quality-testing records traced to Chinese yam powder records",
        "analysis_unit": "Chinese yam powder batch",
        "issue_endpoint": "At least one linked tablet quality-testing record with disintegration time >10 min",
        "issue_count": "33/131 linked Chinese yam powder batches (25.19%)",
        "main_signal": "Higher rejected-material rate and lower 120-mesh fineness",
        "statistical_evidence": "Rejected-material rate: Spearman rho = 0.375 with linked issue rate. Through 120-mesh mean: rho = -0.328 with linked issue rate.",
        "interpretation": "Chinese yam powder records provided a parallel process-material evidence layer for the same finished-product endpoint.",
    },
    {
        "evidence_layer": "Integrated hierarchical diagnosis",
        "linkage_path": "Finished-product issue -> tablet MES -> extract powder / Chenpi / Chinese yam powder",
        "linked_scope": "804 tablet records traced to extract-powder quality data and 839 tablet records traced to Chinese yam powder records",
        "analysis_unit": "Cross-layer batch evidence chain",
        "issue_endpoint": "0.8 g tablet disintegration time >10 min",
        "issue_count": "Layer-dependent linked sample size",
        "main_signal": "Convergent evidence across extract powder, Chenpi, and Chinese yam powder",
        "statistical_evidence": "Extract powder: total ash/extractives. Chenpi: hesperidin/moisture. Chinese yam powder: rejected-material rate/120-mesh fineness.",
        "interpretation": "The issue-driven framework connected the finished-product disintegration issue to upstream material and manufacturing-process risks.",
    },
]


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplementary Table S2"

    ws["A1"] = "Supplementary Table S2. Layer-wise batch-linkage and evidence summary for finished-product disintegration diagnosis"
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    headers = list(rows[0].keys())
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F3A40")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row in enumerate(rows, start=4):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=row[h])
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [26, 42, 52, 28, 34, 32, 30, 48, 52]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 36
    for r in range(4, ws.max_row + 1):
        ws.row_dimensions[r].height = 90

    ws.freeze_panes = "A4"
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
