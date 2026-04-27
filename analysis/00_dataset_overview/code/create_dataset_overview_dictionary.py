from __future__ import annotations

import math
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path.cwd()
OUT_DIR = ROOT / "数据概况表_英文" / "tables"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / "real_world_manufacturing_dataset_overview_and_dictionary.xlsx"


def find_english_data_dir() -> Path:
    for p in ROOT.iterdir():
        if not p.is_dir():
            continue
        names = [x.name for x in p.iterdir() if x.is_file()]
        if any("finished_product_physchem" in n for n in names) and any("yam_powder_mes" in n for n in names):
            return p
    raise FileNotFoundError("Could not locate the finalized English data directory.")


DATA_DIR = find_english_data_dir()

DATASETS = [
    {
        "key": "finished_quality",
        "file_contains": "finished_product_physchem",
        "sheet": "JWXS tablet quality",
        "entity": "Jianwei Xiaoshi tablet quality testing data",
        "layer": "Jianwei Xiaoshi tablet finished-product quality",
        "source": "Quality-control testing records",
        "primary_key": "batch_no",
        "date_col": "production_date",
        "relationship": "Linked to Jianwei Xiaoshi tablet MES production records by finished-product batch identifier.",
        "role": "Describes the finished-product quality records of Jianwei Xiaoshi tablets.",
    },
    {
        "key": "finished_mes",
        "file_contains": "finished_product_mes_main",
        "sheet": "JWXS tablet MES",
        "entity": "Jianwei Xiaoshi tablet MES production records",
        "layer": "Jianwei Xiaoshi tablet manufacturing",
        "source": "MES production records",
        "primary_key": "batch_no",
        "date_col": "production_date",
        "relationship": "Links Jianwei Xiaoshi tablet batches to Jianwei Xiaoshi extract-powder and Chinese yam powder input batches.",
        "role": "Describes the MES production records of Jianwei Xiaoshi tablets and provides batch linkage to material inputs.",
    },
    {
        "key": "extract_quality",
        "file_contains": "extract_powder_physchem",
        "sheet": "JWXS extract powder",
        "entity": "Jianwei Xiaoshi extract-powder quality testing data",
        "layer": "Jianwei Xiaoshi extract-powder quality",
        "source": "Quality-control testing records",
        "primary_key": "batch_no",
        "date_col": "",
        "relationship": "Linked to Jianwei Xiaoshi tablet MES production records by extract-powder batch identifier.",
        "role": "Describes quality testing records of Jianwei Xiaoshi extract powder.",
    },
    {
        "key": "raw_quality",
        "file_contains": "aged_tangerine_peel_test",
        "sheet": "Chenpi quality",
        "entity": "Chenpi quality testing data",
        "layer": "Chenpi raw-material quality",
        "source": "Quality-control testing records",
        "primary_key": "batch_no",
        "date_col": "inspection_date",
        "relationship": "Connected to Jianwei Xiaoshi extract-powder batches through Chenpi batch traceability information.",
        "role": "Describes quality testing records of Chenpi used as a herbal material.",
    },
    {
        "key": "process_material_mes",
        "file_contains": "yam_powder_mes_main",
        "sheet": "Chinese yam powder MES",
        "entity": "Chinese yam powder MES production records",
        "layer": "Chinese yam powder manufacturing",
        "source": "MES production records",
        "primary_key": "batch_no",
        "date_col": "production_date",
        "relationship": "Linked to Jianwei Xiaoshi tablet MES production records by Chinese yam powder batch identifier.",
        "role": "Describes MES production records of Chinese yam powder.",
    },
]


def get_file(file_contains: str) -> Path:
    hits = [p for p in DATA_DIR.iterdir() if p.is_file() and file_contains in p.name and p.suffix.lower() == ".xlsx"]
    if not hits:
        raise FileNotFoundError(file_contains)
    return hits[0]


def normalize_id(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return re.sub(r"\.0$", "", text)


SEP_RE = re.compile(r"[;；,，、\s]+")


def split_values(value) -> list[str]:
    if pd.isna(value):
        return []
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return []
    return [normalize_id(x) for x in SEP_RE.split(text) if x.strip()]


def numeric_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.astype(str).str.replace("%", "", regex=False).str.strip(), errors="coerce")


def flattened_numeric_values(series: pd.Series) -> pd.Series:
    values = []
    for item in series.dropna():
        for part in split_values(item):
            values.append(part)
    return pd.to_numeric(pd.Series(values), errors="coerce").dropna()


def is_multivalue(series: pd.Series, col: str) -> bool:
    if "values" in col:
        return True
    sample = series.dropna().astype(str).head(200)
    if sample.empty:
        return False
    return sample.str.contains(r"[;；,，、]").mean() >= 0.10


def infer_variable_class(df: pd.DataFrame, col: str) -> str:
    lower = col.lower()
    series = df[col]
    if "date" in lower:
        return "time"
    if is_multivalue(series, col):
        return "multi-value numeric"
    if lower.endswith("batch_no") or lower == "batch_no" or "batch_identifier" in lower:
        return "identifier"
    if "supplier" in lower or lower in {"origin", "material_name", "dosage_strength"}:
        return "categorical"
    nums = numeric_series(series)
    non_missing = series.notna().sum()
    if non_missing > 0 and nums.notna().sum() / non_missing >= 0.80:
        return "numeric"
    nunique = series.dropna().astype(str).nunique()
    if nunique <= max(20, len(series) * 0.10):
        return "categorical"
    return "text"


def fmt_num(x, digits=3):
    if x is None or pd.isna(x):
        return ""
    if isinstance(x, (int, float)) and math.isfinite(float(x)):
        return f"{float(x):.{digits}f}".rstrip("0").rstrip(".")
    return str(x)


def contains_cjk(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(text)))


def summarize_column(df: pd.DataFrame, col: str) -> dict:
    s = df[col]
    n = len(df)
    non_missing = int(s.notna().sum())
    missing = n - non_missing
    missing_pct = missing / n * 100 if n else 0
    unique_non_missing = int(s.dropna().astype(str).nunique())
    vclass = infer_variable_class(df, col)

    out = {
        "column_name": col,
        "variable_class": vclass,
        "non_missing_count": non_missing,
        "missing_count": missing,
        "missing_pct": round(missing_pct, 2),
        "unique_non_missing_values": unique_non_missing,
        "minimum": "",
        "maximum": "",
        "mean": "",
        "median": "",
        "date_min": "",
        "date_max": "",
        "category_or_value_summary": "",
        "notes": "",
    }

    if vclass == "time":
        dt = pd.to_datetime(s, errors="coerce")
        out["date_min"] = dt.min().date().isoformat() if dt.notna().any() else ""
        out["date_max"] = dt.max().date().isoformat() if dt.notna().any() else ""
        out["notes"] = "Calendar date variable."
    elif vclass == "numeric":
        nums = numeric_series(s).dropna()
        if not nums.empty:
            out["minimum"] = fmt_num(nums.min())
            out["maximum"] = fmt_num(nums.max())
            out["mean"] = fmt_num(nums.mean())
            out["median"] = fmt_num(nums.median())
        out["notes"] = "Single numeric measurement per record."
    elif vclass == "multi-value numeric":
        nums = flattened_numeric_values(s)
        lengths = s.dropna().map(lambda x: len(split_values(x)))
        if not nums.empty:
            out["minimum"] = fmt_num(nums.min())
            out["maximum"] = fmt_num(nums.max())
            out["mean"] = fmt_num(nums.mean())
            out["median"] = fmt_num(nums.median())
        if not lengths.empty:
            out["category_or_value_summary"] = f"Values per record: {int(lengths.min())}-{int(lengths.max())}; flattened numeric values: {len(nums)}"
        out["notes"] = "Multiple within-batch measurements stored in one cell; use summary variables for primary analysis when available."
    elif vclass == "categorical":
        vc = s.dropna().astype(str).value_counts().head(8)
        all_values = s.dropna().astype(str).tolist()
        if any(contains_cjk(v) for v in all_values):
            total_categories = s.dropna().astype(str).nunique()
            top_count = int(s.dropna().astype(str).value_counts().iloc[0]) if total_categories else 0
            out["category_or_value_summary"] = f"{total_categories} categories; largest category count: {top_count}"
        else:
            out["category_or_value_summary"] = "; ".join([f"{idx} ({cnt})" for idx, cnt in vc.items()])
        out["notes"] = "Categorical descriptor."
    elif vclass == "identifier":
        examples = [x for x in s.dropna().astype(str).head(5).tolist() if not contains_cjk(x)]
        out["category_or_value_summary"] = "Examples: " + ", ".join(examples) if examples else f"{unique_non_missing} unique identifiers"
        out["notes"] = "Batch or material identifier used for linkage; not treated as a continuous variable."
    else:
        examples = [x for x in s.dropna().astype(str).head(5).tolist() if not contains_cjk(x)]
        out["category_or_value_summary"] = "Examples: " + ", ".join(examples) if examples else f"{unique_non_missing} unique text values"
        out["notes"] = "Free-text or high-cardinality descriptor."
    return out


def dataset_date_range(df: pd.DataFrame, date_col: str) -> str:
    if not date_col or date_col not in df.columns:
        return "Not date-indexed"
    dt = pd.to_datetime(df[date_col], errors="coerce")
    if not dt.notna().any():
        return "Not available"
    return f"{dt.min().date().isoformat()} to {dt.max().date().isoformat()}"


def load_all() -> dict[str, pd.DataFrame]:
    loaded = {}
    for meta in DATASETS:
        loaded[meta["key"]] = pd.read_excel(get_file(meta["file_contains"]), dtype=object)
    return loaded


def compute_linkages(data: dict[str, pd.DataFrame]) -> dict[str, str]:
    d2 = data["finished_quality"]
    d3 = data["finished_mes"]
    d4 = data["extract_quality"]
    d6 = data["raw_quality"]
    d7 = data["process_material_mes"]

    d2_batches = set(d2["batch_no"].dropna().map(normalize_id))
    d3_batches = set(d3["batch_no"].dropna().map(normalize_id))
    d4_batches = set(d4["batch_no"].dropna().map(normalize_id))
    d6_batches = set(d6["batch_no"].dropna().map(normalize_id))
    d7_batches = set(d7["batch_no"].dropna().map(normalize_id))

    d2_rows_link_d3 = sum(normalize_id(x) in d3_batches for x in d2["batch_no"])
    d2_batches_link_d3 = len(d2_batches & d3_batches)

    d3_extract_rows = 0
    d3_yam_rows = 0
    d4_used = set()
    d7_used = set()
    d3_extract_map = {}
    d3_yam_map = {}
    for _, row in d3.iterrows():
        batch = normalize_id(row["batch_no"])
        eps = [x for x in split_values(row.get("jwxs_extract_powder_batch_no", "")) if x in d4_batches]
        yps = [x for x in split_values(row.get("yam_powder_luoting_batch_no", "")) if x in d7_batches]
        if eps:
            d3_extract_rows += 1
            d4_used.update(eps)
        if yps:
            d3_yam_rows += 1
            d7_used.update(yps)
        d3_extract_map[batch] = bool(eps)
        d3_yam_map[batch] = bool(yps)

    d2_extract_rows = sum(d3_extract_map.get(normalize_id(x), False) for x in d2["batch_no"])
    d2_yam_rows = sum(d3_yam_map.get(normalize_id(x), False) for x in d2["batch_no"])

    # Traceability bridge is not a formal dataset in the current data dictionary, but is used only for this linkage count.
    trace_file = None
    for p in DATA_DIR.iterdir():
        if p.is_file() and "extract_powder_traceability" in p.name and p.suffix.lower() == ".xlsx":
            trace_file = p
            break
    raw_bridge = "Not applicable"
    if trace_file is not None:
        trace = pd.read_excel(trace_file, dtype=object)
        chenpi_label = "\u9648\u76ae"
        chenpi = trace[trace["material_type"].astype(str) == chenpi_label] if "material_type" in trace.columns else trace.iloc[0:0]
        if not chenpi.empty:
            raw_batches = set(chenpi["material_batch_no"].dropna().map(normalize_id))
            matched_raw = raw_batches & d6_batches
            matched_rows = chenpi[chenpi["material_batch_no"].map(normalize_id).isin(matched_raw)]
            matched_extract = set(matched_rows["extract_powder_batch_no"].dropna().map(normalize_id))
            raw_bridge = f"{len(matched_raw)} Chenpi testing batches linked; {len(matched_extract & d4_batches)} Jianwei Xiaoshi extract-powder testing batches connected"

    return {
        "finished_quality": f"{d2_rows_link_d3} Jianwei Xiaoshi tablet quality-testing records linked to tablet MES records by finished-product batch identifier ({d2_batches_link_d3} distinct batch identifiers).",
        "finished_mes": f"{d3_extract_rows} tablet MES records linked to Jianwei Xiaoshi extract-powder quality data; {d3_yam_rows} tablet MES records linked to Chinese yam powder MES data.",
        "extract_quality": f"{len(d4_used)} Jianwei Xiaoshi extract-powder batches used in tablet MES records; {d2_extract_rows} tablet quality-testing records linked end-to-end through MES.",
        "raw_quality": raw_bridge,
        "process_material_mes": f"{len(d7_used)} Chinese yam powder batches used in tablet MES records; {d2_yam_rows} tablet quality-testing records linked end-to-end through MES.",
    }


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for j, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(start_row, j, col_name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F3A40")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            cell = ws.cell(i, j, "" if pd.isna(value) else value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_sheet(ws, freeze: str = "A2"):
    ws.freeze_panes = freeze
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            text = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, min(len(text), 70))
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, 48))
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 28 if row_idx == 1 else 42


def build_workbook():
    data = load_all()
    linkages = compute_linkages(data)

    overview_rows = []
    column_summaries = {}
    for meta in DATASETS:
        df = data[meta["key"]]
        summary = pd.DataFrame([summarize_column(df, c) for c in df.columns])
        column_summaries[meta["key"]] = summary
        class_counts = summary["variable_class"].value_counts().to_dict()
        missing_cells = int(summary["missing_count"].sum())
        total_cells = int(len(df) * len(df.columns))
        overview_rows.append({
            "dataset_entity": meta["entity"],
            "manufacturing_layer": meta["layer"],
            "source_type": meta["source"],
            "records": len(df),
            "variables": len(df.columns),
            "identifier_variables": class_counts.get("identifier", 0),
            "time_variables": class_counts.get("time", 0),
            "numeric_variables": class_counts.get("numeric", 0),
            "multi_value_numeric_variables": class_counts.get("multi-value numeric", 0),
            "categorical_variables": class_counts.get("categorical", 0),
            "text_variables": class_counts.get("text", 0),
            "missing_cells": missing_cells,
            "missing_cell_pct": round(missing_cells / total_cells * 100, 2) if total_cells else 0,
            "primary_linkage_key": meta["primary_key"],
            "date_range": dataset_date_range(df, meta["date_col"]),
            "batch-level_relationship": meta["relationship"],
            "observed_linkage_coverage": linkages[meta["key"]],
            "analytical_role": meta["role"],
        })
    overview = pd.DataFrame(overview_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset overview"
    write_dataframe(ws, overview)
    style_sheet(ws)

    for meta in DATASETS:
        summary = column_summaries[meta["key"]]
        ws = wb.create_sheet(meta["sheet"][:31])
        write_dataframe(ws, summary)
        style_sheet(ws)

    wb.save(OUT_FILE)
    return OUT_FILE


if __name__ == "__main__":
    out = build_workbook()
    print(out)
